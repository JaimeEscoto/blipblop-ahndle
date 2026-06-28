"""Genera las 3 plantillas Excel del piloto OdontiaCloud."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

BRAND_DARK = "0F4A3F"      # verde oscuro odontiacloud
BRAND_TEAL = "2EA8A0"      # teal claro
HEADER_FILL = PatternFill("solid", start_color=BRAND_DARK)
SUBHEADER_FILL = PatternFill("solid", start_color="E6F4F1")
EXAMPLE_FILL = PatternFill("solid", start_color="FFF8E1")
THIN = Side(border_style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=BRAND_DARK)
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="546E7A")
BODY_FONT = Font(name="Arial", size=11, color="263238")
EXAMPLE_FONT = Font(name="Arial", size=10, italic=True, color="6D4C00")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_header(ws, headers, row=4):
    for idx, (col_title, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=col_title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        if comment:
            cell.comment = Comment(comment, "OdontiaCloud")
    ws.row_dimensions[row].height = 32


def add_title(ws, title, subtitle):
    ws["A1"] = "OdontiaCloud — Piloto Clínicas"
    ws["A1"].font = Font(name="Arial", size=12, bold=True, color=BRAND_TEAL)
    ws["A2"] = title
    ws["A2"].font = TITLE_FONT
    ws["A3"] = subtitle
    ws["A3"].font = NOTE_FONT
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18


def add_example_row(ws, values, row):
    for idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=idx, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.alignment = LEFT
        cell.border = BORDER
    ws.cell(row=row, column=len(values)+1, value="EJEMPLO — eliminar antes de cargar").font = NOTE_FONT


def set_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ---------- 1. PACIENTES ----------
wb = Workbook()
ws = wb.active
ws.title = "Pacientes"

add_title(
    ws,
    "Plantilla de Pacientes",
    "Llena una fila por paciente. Los campos marcados con * son obligatorios.",
)

headers = [
    ("Nombre completo *", "Nombres y apellidos del paciente. Ej: María Fernanda López"),
    ("Email *", "Correo único por paciente. Será su usuario en OdontiaCloud."),
    ("Teléfono", "Incluye el indicativo. Ej: +57 311 234 5678"),
    ("Documento de identidad", "Cédula, DNI o pasaporte. Solo dígitos o letras, sin puntos."),
    ("Tipo de sangre", "Opcional. Ej: O+, A-, AB+"),
    ("Alergias", "Opcional. Separar varias con coma. Ej: penicilina, látex"),
    ("Condiciones médicas", "Opcional. Ej: diabetes tipo 2, hipertensión"),
    ("Medicamentos actuales", "Opcional. Ej: losartán 50mg, metformina 850mg"),
    ("Contacto de emergencia", "Opcional. Nombre completo de la persona a contactar."),
    ("Teléfono de emergencia", "Opcional. Ej: +57 320 555 1212"),
]
apply_header(ws, headers, row=4)

add_example_row(
    ws,
    [
        "María Fernanda López",
        "maria.lopez@gmail.com",
        "+57 320 345 6789",
        "1087654321",
        "O+",
        "penicilina",
        "hipertensión",
        "losartán 50mg",
        "Carlos López (esposo)",
        "+57 311 999 0011",
    ],
    row=5,
)

set_widths(
    ws,
    {
        "A": 28, "B": 30, "C": 20, "D": 22, "E": 14,
        "F": 24, "G": 26, "H": 26, "I": 26, "J": 22, "K": 36,
    },
)
ws.freeze_panes = "A5"


# ---------- 2. DOCTORES ----------
ws2 = wb.create_sheet("Doctores")
add_title(
    ws2,
    "Plantilla de Doctores",
    "Una fila por profesional. El email debe ser único (será su acceso a OdontiaCloud).",
)
headers2 = [
    ("Nombre completo *", "Nombre del doctor con título. Ej: Dr. Andrés Pardo"),
    ("Especialidad *", "Ej: Odontología general, Ortodoncia, Endodoncia, Cirugía oral"),
    ("Email *", "Correo único — será su usuario para iniciar sesión."),
    ("Teléfono", "Opcional. Ej: +57 318 111 2233"),
    ("Número de licencia / RM", "Registro médico u odontológico. Opcional pero recomendado."),
]
apply_header(ws2, headers2, row=4)
add_example_row(
    ws2,
    [
        "Dr. Andrés Pardo",
        "Ortodoncia",
        "andres.pardo@clinicasonrisa.com",
        "+57 318 111 2233",
        "RM-OD-45821",
    ],
    row=5,
)
set_widths(ws2, {"A": 28, "B": 24, "C": 32, "D": 20, "E": 22, "F": 36})
ws2.freeze_panes = "A5"


# ---------- 3. CITAS ----------
ws3 = wb.create_sheet("Citas")
add_title(
    ws3,
    "Plantilla de Citas",
    "Una fila por cita. Pacientes y doctores se vinculan por email — deben existir en sus plantillas.",
)
headers3 = [
    ("Email del paciente *", "Debe coincidir con un email cargado en la hoja Pacientes."),
    ("Email del doctor *", "Debe coincidir con un email cargado en la hoja Doctores."),
    ("Fecha *", "Formato AAAA-MM-DD. Ej: 2026-07-15"),
    ("Hora *", "Formato 24h HH:MM. Ej: 09:30"),
    ("Motivo", "Breve. Ej: Control de ortodoncia, Limpieza"),
    ("Estado *", "scheduled (programada), completed (realizada) o cancelled (cancelada)"),
    ("Notas", "Observaciones internas, opcional."),
]
apply_header(ws3, headers3, row=4)
add_example_row(
    ws3,
    [
        "maria.lopez@gmail.com",
        "andres.pardo@clinicasonrisa.com",
        "2026-07-15",
        "09:30",
        "Control de ortodoncia",
        "scheduled",
        "Traer radiografía panorámica reciente",
    ],
    row=5,
)
set_widths(ws3, {"A": 30, "B": 32, "C": 14, "D": 10, "E": 28, "F": 14, "G": 32, "H": 36})

# Lista desplegable para Estado
dv = DataValidation(type="list", formula1='"scheduled,completed,cancelled"', allow_blank=False)
dv.error = "Solo se permite: scheduled, completed o cancelled"
dv.errorTitle = "Estado inválido"
ws3.add_data_validation(dv)
dv.add("F6:F1000")

ws3.freeze_panes = "A5"


# ---------- 4. PROCEDIMIENTOS ----------
ws4 = wb.create_sheet("Procedimientos")
add_title(
    ws4,
    "Catálogo de Procedimientos",
    "Una fila por servicio que ofrece la clínica. Se usa para cotizar y facturar.",
)
headers4 = [
    ("Código", "Opcional. Código interno o CUPS. Ej: D-001, 997102"),
    ("Nombre *", "Nombre del procedimiento. Ej: Profilaxis dental"),
    ("Descripción", "Detalle breve para el paciente. Opcional."),
    ("Precio por defecto *", "Solo número, sin símbolos. Ej: 80000 (no $80.000)"),
    ("Duración (minutos)", "Tiempo estimado en minutos. Ej: 30"),
    ("Activo *", "SI = visible para cobrar / NO = oculto del catálogo"),
]
apply_header(ws4, headers4, row=4)

ejemplos_proc = [
    ["D-001", "Consulta odontológica general", "Valoración inicial y diagnóstico", 50000, 20, "SI"],
    ["D-002", "Profilaxis dental", "Limpieza y pulido dental completo", 80000, 30, "SI"],
    ["D-003", "Resina de fotocurado simple", "Restauración estética de una cara", 120000, 45, "SI"],
    ["D-004", "Extracción dental simple", "Exodoncia de pieza erupcionada", 90000, 30, "SI"],
    ["D-005", "Endodoncia unirradicular", "Tratamiento de conducto en pieza anterior", 350000, 60, "SI"],
    ["D-006", "Radiografía periapical", "Imagen diagnóstica individual", 25000, 10, "SI"],
]
for i, vals in enumerate(ejemplos_proc):
    row = 5 + i
    for idx, val in enumerate(vals, start=1):
        cell = ws4.cell(row=row, column=idx, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.alignment = LEFT
        cell.border = BORDER
        if idx == 4:  # precio
            cell.number_format = '#,##0'
    ws4.cell(row=row, column=len(vals)+1, value="EJEMPLO — eliminar o reemplazar").font = NOTE_FONT

set_widths(ws4, {"A": 12, "B": 32, "C": 36, "D": 18, "E": 14, "F": 10, "G": 32})

# Validación de "Activo"
dv_act = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
dv_act.error = "Solo se permite SI o NO"
dv_act.errorTitle = "Valor inválido"
ws4.add_data_validation(dv_act)
dv_act.add("F5:F1000")

ws4.freeze_panes = "A5"


# ---------- Hoja de instrucciones rápidas ----------
ws0 = wb.create_sheet("Instrucciones", 0)
ws0["A1"] = "OdontiaCloud — Carga inicial del piloto"
ws0["A1"].font = Font(name="Arial", size=16, bold=True, color=BRAND_DARK)
ws0["A2"] = "Sigue estos 4 pasos para que tu clínica quede lista en el sistema."
ws0["A2"].font = NOTE_FONT

steps = [
    ("1. Pacientes", "Llena la hoja 'Pacientes' con todos los pacientes activos de la clínica. Email obligatorio y único."),
    ("2. Doctores", "Llena la hoja 'Doctores' con los profesionales que atenderán en la clínica. Email obligatorio y único."),
    ("3. Procedimientos", "Llena la hoja 'Procedimientos' con el catálogo de servicios y precios que ofrece la clínica."),
    ("4. Citas", "Llena la hoja 'Citas' con la agenda del próximo mes. Los emails deben coincidir con los de las hojas anteriores."),
    ("5. Envío", "Borra las filas de ejemplo (en amarillo) y envía este archivo al equipo OdontiaCloud para la importación."),
]
for i, (title, desc) in enumerate(steps, start=4):
    ws0.cell(row=i, column=1, value=title).font = Font(name="Arial", size=12, bold=True, color=BRAND_TEAL)
    ws0.cell(row=i, column=2, value=desc).font = BODY_FONT
    ws0.cell(row=i, column=2).alignment = LEFT
    ws0.row_dimensions[i].height = 32

ws0["A10"] = "Soporte"
ws0["A10"].font = Font(name="Arial", size=12, bold=True, color=BRAND_DARK)
ws0["B10"] = "Si tienes dudas, escríbenos antes de enviar el archivo."
ws0["B10"].font = BODY_FONT

set_widths(ws0, {"A": 18, "B": 90})

# Reordenar: Instrucciones, Pacientes, Doctores, Procedimientos, Citas
wb.move_sheet("Procedimientos", offset=-1)

wb.save("/Users/maletincorp/Documents/insigne-pro/piloto_clinicas/OdontiaCloud_Plantilla_Carga_Inicial.xlsx")
print("OK plantilla generada")
